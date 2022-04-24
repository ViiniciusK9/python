'''
https://openpyxl.readthedocs.io/en/stable/
pip install openpyxl
'''
import openpyxl
from random import uniform

pedidos = openpyxl.load_workbook('pedidos.xlsx')

nome_planilhas = pedidos.sheetnames
print(nome_planilhas, end='\n\n')
planilha_0 = pedidos[nome_planilhas[0]]

# Selecionar um campo especifico
campo = planilha_0['b1']
print(f'Campo[b1]: {campo.value}', end='\n\n')

# Selecionar uma coluna
coluna = planilha_0['b']
print(f'Coluna[b]: {[x.value for x in coluna if x.value is not None]}', end='\n\n')

# selecionar uma linha 
linha = planilha_0['2']
print(f'Linha: {[x.value for x in linha if x.value is not None]}', end='\n\n')

# Percorrer a planilha toda.
for linha in planilha_0:
    for coluna in linha:
        if coluna.value is not None:
            print(coluna.value, end='\t')
    print()


# Editar um campo específico.
planilha_0['b3'].value = 4000

# Salvando em um novo arquivo as alterações feitas.
pedidos.save('novo_arquivo_editado.xlsx')

# Percorendo e editando a planilha.
for linha in range(5, 16):
    planilha_0.cell(linha, 1).value = (linha - 1)
    planilha_0.cell(linha, 2).value = (1200 + linha)
    planilha_0.cell(linha, 3).value = f'R$ {uniform(10, 1000):.2f}'

pedidos.save('novo_arquivo_editado_2.xlsx')


# Criando uma nova planilha.
planilha = openpyxl.Workbook()

# Criando planilhas no arquivo .xlsx
planilha.create_sheet('Planilha1', 0)
planilha.create_sheet('Planilha2', 1)

planilha_1 = planilha['Planilha1']
planilha_2 = planilha['Planilha2']

# Preenchendo as celulas com dados aleatorios.
for linha in range(1, 11):
    for coluna in range(1, 6):
        planilha_1.cell(linha, coluna).value = f'{uniform(10, 50):.2f}'
        planilha_2.cell(linha, coluna).value = f'{uniform(10, 50):.2f}'

# Salvando a nova planilha.
planilha.save('random_planilha.xlsx')
